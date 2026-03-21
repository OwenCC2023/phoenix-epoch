#!/usr/bin/env python3
"""
Generate effects_matrix.xlsx — a full cross-reference of every game element
(building, government option, trait) against every effect category.

Run from the repo root:
    python generate_effects_matrix.py

Output: effects_matrix.xlsx in the current directory.
"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "backend"))

from provinces.building_constants import BUILDING_TYPES
from nations.government_constants import (
    GOV_DIRECTION, GOV_ECONOMIC_CATEGORY, GOV_STRUCTURE,
    GOV_POWER_ORIGIN, GOV_POWER_TYPE,
)
from nations.trait_constants import TRAIT_DEFS

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
C_SHEET_HDR   = "1F3864"   # dark navy     — sheet/axis header
C_GROUP_HDR   = "2E75B6"   # mid blue      — column-group header
C_COL_HDR     = "BDD7EE"   # pale blue     — individual column header
C_ROW_CAT     = "D6E4F7"   # lighter blue  — source-category rows (Buildings, etc.)
C_POS         = "C6EFCE"   # green fill    — positive value
C_NEG         = "FFC7CE"   # red fill      — negative value
C_STUB        = "FFF2CC"   # yellow fill   — stub/not-yet-wired value
C_ROW_A       = "F9F9F9"   # alternating row A
C_ROW_B       = "FFFFFF"   # alternating row B

def fill(hex_str):
    return PatternFill("solid", fgColor=hex_str)

def bold(color="000000", size=10):
    return Font(bold=True, color=color, size=size)

def std(size=10):
    return Font(size=size)

def center(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ---------------------------------------------------------------------------
# Column definitions
# Three kinds of values:
#   • numeric  — shown as percentage or flat number, green/red fill
#   • capacity — shown as integer, green fill only
#   • text     — shown as-is (building_restrictions list)
#   • stub     — yellow fill, not yet wired into simulation
# ---------------------------------------------------------------------------

# Each entry: (group_label, col_key, short_label, value_type)
# value_type: "pct" | "flat" | "cap" | "text" | "stub_pct" | "stub_flat"
COLUMN_DEFS = [
    # ---- Province-scope building effects ----
    ("Province Effects", "stability_recovery_bonus",  "+Stability\nRecovery/turn",  "pct"),
    ("Province Effects", "growth_bonus",              "+Growth\n/turn",             "pct"),
    ("Province Effects", "farming_bonus",             "+Farming\nBonus",            "pct"),
    ("Province Effects", "research_bonus",            "+Research\nBonus",           "pct"),
    ("Province Effects", "integration_bonus",         "+Integration\nBonus",        "pct"),
    ("Province Effects", "construction_time_reduction","−Construction\nTime",       "pct"),
    ("Province Effects", "literacy_bonus",            "+Literacy\nBonus",           "pct"),
    ("Province Effects", "march_speed_bonus",         "+March\nSpeed",              "pct"),
    ("Province Effects", "sea_transit_speed",         "+Sea\nTransit",              "pct"),
    ("Province Effects", "river_transit_speed",       "+River\nTransit",            "pct"),
    ("Province Effects", "air_transit_speed",         "+Air\nTransit",              "pct"),

    # ---- National capacity (use-it-or-lose-it) ----
    ("National Capacity", "land_trade_capacity",      "Land Trade\nCapacity",       "cap"),
    ("National Capacity", "naval_trade_capacity",     "Naval Trade\nCapacity",      "cap"),
    ("National Capacity", "air_trade_capacity",       "Air Trade\nCapacity",        "cap"),
    ("National Capacity", "bureaucratic_capacity",    "Bureaucratic\nCapacity",     "cap"),
    ("National Capacity", "upkeep_reduction",         "−Upkeep",                    "pct"),
    ("National Capacity", "construction_cost_reduction", "−Build\nCost",            "pct"),

    # ---- Military building effects (stubs) ----
    ("Military (stub)", "army_training_speed_bonus",  "+Army\nTraining",            "stub_pct"),
    ("Military (stub)", "army_combat_bonus",          "+Army\nCombat",              "stub_pct"),
    ("Military (stub)", "army_upkeep_reduction",      "−Army\nUpkeep",              "stub_pct"),
    ("Military (stub)", "navy_training_speed_bonus",  "+Navy\nTraining",            "stub_pct"),
    ("Military (stub)", "navy_combat_bonus",          "+Navy\nCombat",              "stub_pct"),
    ("Military (stub)", "navy_upkeep_reduction",      "−Navy\nUpkeep",              "stub_pct"),
    ("Military (stub)", "air_training_speed_bonus",   "+Air\nTraining",             "stub_pct"),
    ("Military (stub)", "air_combat_bonus",           "+Air\nCombat",               "stub_pct"),
    ("Military (stub)", "air_upkeep_reduction",       "−Air\nUpkeep",               "stub_pct"),

    # ---- Government / national modifiers ----
    ("Gov & Trait Modifiers", "stability",            "Stability\n(flat)",          "flat"),
    ("Gov & Trait Modifiers", "growth",               "Growth\n/turn",              "pct"),
    ("Gov & Trait Modifiers", "integration",          "Integration\n(%)",           "pct"),
    ("Gov & Trait Modifiers", "trade",                "Trade\n(%)",                 "pct"),
    ("Gov & Trait Modifiers", "research",             "Research\n(%)",              "pct"),
    ("Gov & Trait Modifiers", "military",             "Military\n(%)",              "pct"),
    ("Gov & Trait Modifiers", "consumption",          "Consumption\n(%)",           "pct"),
    ("Gov & Trait Modifiers", "prod_food",            "Production\nFood (%)",       "pct"),
    ("Gov & Trait Modifiers", "prod_materials",       "Production\nMaterials (%)",  "pct"),
    ("Gov & Trait Modifiers", "prod_wealth",          "Production\nWealth (%)",     "pct"),
    ("Gov & Trait Modifiers", "prod_energy",          "Production\nEnergy (%)",     "pct"),
    ("Gov & Trait Modifiers", "prod_manpower",        "Production\nManpower (%)",   "pct"),

    # ---- Building efficiency by category ----
    ("Building Efficiency", "be_financial",              "BE: Financial",              "pct"),
    ("Building Efficiency", "be_light_manufacturing",    "BE: Light\nManuf.",          "pct"),
    ("Building Efficiency", "be_heavy_manufacturing",    "BE: Heavy\nManuf.",          "pct"),
    ("Building Efficiency", "be_refining",               "BE: Refining",               "pct"),
    ("Building Efficiency", "be_chemical",               "BE: Chemical",               "pct"),
    ("Building Efficiency", "be_pharmaceutical",         "BE: Pharma",                 "pct"),
    ("Building Efficiency", "be_farming",                "BE: Farming",                "pct"),
    ("Building Efficiency", "be_extraction",             "BE: Extraction",             "pct"),
    ("Building Efficiency", "be_construction",           "BE: Construction",           "pct"),
    ("Building Efficiency", "be_transport",              "BE: Transport",              "pct"),
    ("Building Efficiency", "be_communications",         "BE: Comms",                  "pct"),
    ("Building Efficiency", "be_entertainment",          "BE: Entertainment",          "pct"),
    ("Building Efficiency", "be_healthcare",             "BE: Healthcare",             "pct"),
    ("Building Efficiency", "be_religious",              "BE: Religious",              "pct"),
    ("Building Efficiency", "be_green_energy",           "BE: Green\nEnergy",          "pct"),
    ("Building Efficiency", "be_government_regulatory",  "BE: Gov\nRegulatory",        "pct"),
    ("Building Efficiency", "be_government_oversight",   "BE: Gov\nOversight",         "pct"),
    ("Building Efficiency", "be_government_management",  "BE: Gov\nManagement",        "pct"),
    ("Building Efficiency", "be_government_security",    "BE: Gov\nSecurity",          "pct"),
    ("Building Efficiency", "be_government_education",   "BE: Gov\nEducation",         "pct"),
    ("Building Efficiency", "be_government_organization","BE: Gov\nOrganization",      "pct"),
    ("Building Efficiency", "be_government_welfare",     "BE: Gov\nWelfare",           "pct"),
    ("Building Efficiency", "be_military_education",     "BE: Military\nEducation",    "pct"),

    # ---- Trait-specific effects ----
    ("Trait Effects", "manpower_bonus",               "+Manpower\n(%)",             "pct"),
    ("Trait Effects", "wealth_production_bonus",      "+Wealth\nProd (%)",          "pct"),
    ("Trait Effects", "food_production_bonus",        "+Food\nProd (%)",            "pct"),
    ("Trait Effects", "rural_output_bonus",           "+Rural\nOutput",             "pct"),
    ("Trait Effects", "rural_output_penalty",         "−Rural\nOutput",             "pct"),
    ("Trait Effects", "urban_output_bonus",           "+Urban\nOutput",             "pct"),
    ("Trait Effects", "urban_growth_penalty",         "−Urban\nGrowth/turn",        "pct"),
    ("Trait Effects", "urban_threshold_reduction",    "Urban\nThreshold −",         "flat"),
    ("Trait Effects", "training_speed_bonus",         "+Training\nSpeed",           "stub_pct"),
    ("Trait Effects", "military_upkeep_reduction",    "−Mil.\nUpkeep",              "stub_pct"),
    ("Trait Effects", "building_restrictions",        "Building\nRestrictions",     "text"),

    # ---- Stub effects (future systems) ----
    ("Stubs", "trade_capacity_bonus",        "Trade Cap\nBonus (stub)",     "stub_pct"),
    ("Stubs", "trade_capacity_penalty",      "Trade Cap\nPenalty (stub)",   "stub_pct"),
    ("Stubs", "diplomatic_reputation_bonus", "Diplo Rep\n+ (stub)",         "stub_pct"),
    ("Stubs", "diplomatic_reputation_penalty","Diplo Rep\n− (stub)",        "stub_pct"),
    ("Stubs", "espionage_effectiveness",     "Espionage\n(stub)",           "stub_pct"),
    ("Stubs", "counter_espionage",           "Counter\nEspionage (stub)",   "stub_pct"),
    ("Stubs", "arms_production_bonus",       "Arms Prod\nBonus (stub)",     "stub_pct"),
    ("Stubs", "arms_production_penalty",     "Arms Prod\nPenalty (stub)",   "stub_pct"),
]

COL_KEYS = [cd[1] for cd in COLUMN_DEFS]


# ---------------------------------------------------------------------------
# Data extraction helpers
# ---------------------------------------------------------------------------

def flatten_building_effects(raw_effects: dict) -> dict:
    """Convert a raw base_effects dict to flat col_key → value."""
    out = {}
    for k, v in raw_effects.items():
        out[k] = v          # all building effects map directly to col_key
    return out


def flatten_gov_effects(raw: dict) -> dict:
    """Flatten a government component effects dict to col_key → value."""
    out = {}
    for k, v in raw.items():
        if k == "building_efficiency":
            for cat, bonus in v.items():
                out[f"be_{cat}"] = bonus
        elif k == "production":
            for res, bonus in v.items():
                out[f"prod_{res}"] = bonus
        elif k in ("policy_effectiveness", "military_effectiveness"):
            pass    # stubs not shown in this matrix
        else:
            out[k] = v
    return out


def flatten_trait_effects(raw: dict) -> dict:
    """Flatten a trait strong_effects or weak_effects dict to col_key → value."""
    out = {}
    for k, v in raw.items():
        if k == "building_efficiency_bonus":
            for cat, bonus in v.items():
                out[f"be_{cat}"] = bonus
        elif k in ("policy_constraints", "technology_adoption_speed",
                   "technology_adoption_penalty", "policy_change_speed",
                   "minority_policy_expectation", "policy_change_resistance",
                   "treaty_bureaucratic_reduction", "treaty_break_cost_reduction",
                   "entrepreneurship_bonus", "bureaucratic_capacity_bonus",
                   "urban_migration_bonus", "urban_emigration_bonus",
                   "happiness_baseline_bonus", "happiness_baseline_penalty",
                   "government_building_cost_reduction",
                   "government_building_cost_increase",
                   "elite_institution_penalty",
                   "military_organisation_bonus",
                   "literacy_bonus", "literacy_penalty",
                   "conscription_penalty"):
            pass    # skip system stubs not in column list
        elif k == "research_penalty":
            out["research"] = v         # normalise to shared "research" col
        elif k == "growth_penalty":
            out["growth"] = v
        elif k == "stability_penalty":
            out["stability"] = v
        elif k == "stability_bonus":
            out["stability"] = v
        elif k == "integration_bonus":
            out["integration"] = v
        else:
            out[k] = v
    return out


# ---------------------------------------------------------------------------
# Build row data
# ---------------------------------------------------------------------------

def building_rows():
    """Yield (source_type, label, category, note, flat_effects) for buildings."""
    # Canonical category order from CLAUDE.md
    CAT_ORDER = [
        "financial", "light_manufacturing", "heavy_manufacturing", "refining",
        "chemical", "pharmaceutical", "farming", "extraction", "construction",
        "transport", "communications", "entertainment", "healthcare", "religious",
        "green_energy", "government_regulatory", "government_oversight",
        "government_management", "government_security", "government_education",
        "government_organization", "government_welfare", "military_education",
    ]
    cat_rank = {c: i for i, c in enumerate(CAT_ORDER)}

    buildings = sorted(
        BUILDING_TYPES.items(),
        key=lambda x: (cat_rank.get(x[1].get("category", ""), 99), x[0]),
    )

    last_cat = None
    for btype, bdata in buildings:
        cat = bdata.get("category", "unknown")
        effects = flatten_building_effects(bdata.get("base_effects", {}))
        note = "L1 base; scales ×(1 + 0.9×log₁₀(N)) per level"
        yield ("Building", bdata.get("label", btype), cat, note, effects, last_cat != cat)
        last_cat = cat


def government_rows():
    """Yield (source_type, label, axis, note, flat_effects, is_first_in_group) for gov options."""
    axes = [
        ("Direction",      GOV_DIRECTION),
        ("Economic Cat.",  GOV_ECONOMIC_CATEGORY),
        ("Structure",      GOV_STRUCTURE),
        ("Power Origin",   GOV_POWER_ORIGIN),
        ("Power Type",     GOV_POWER_TYPE),
    ]
    LABELS = {
        "top_down": "Top-Down", "bottom_up": "Bottom-Up", "none": "None",
        "liberal": "Liberal", "collectivist": "Collectivist",
        "protectionist": "Protectionist", "resource": "Resource",
        "autarkic": "Autarkic", "subsistence": "Subsistence",
        "hereditary": "Hereditary", "power_consensus": "Power-Consensus",
        "federal": "Federal", "representative": "Representative", "direct": "Direct",
        "elections": "Elections", "economic_success": "Economic Success",
        "law_and_order": "Law & Order", "military_power": "Military Power",
        "religious": "Religious", "ideology": "Ideology",
        "singular": "Singular", "council": "Council", "large_body": "Large Body",
        "multi_body": "Multi-body", "staggered_groups": "Staggered Groups",
    }
    for axis_name, axis_dict in axes:
        first = True
        for key, raw in axis_dict.items():
            effects = flatten_gov_effects(raw)
            yield ("Government", LABELS.get(key, key), axis_name, "per-component (not combined)", effects, first)
            first = False


def trait_rows():
    """Yield (source_type, label, pair, note, flat_effects, is_first_in_group) for traits."""
    for tkey, tdata in TRAIT_DEFS.items():
        pair_idx = tdata["pair_index"]
        for strength, effects_key in (("Strong", "strong_effects"), ("Weak", "weak_effects")):
            raw = tdata.get(effects_key, {})
            effects = flatten_trait_effects(raw)
            label = f"{tdata['name']} ({strength})"
            group = f"Pair {pair_idx}: {tkey}"
            is_first = strength == "Strong"
            yield ("Trait", label, group, strength, effects, is_first)


# ---------------------------------------------------------------------------
# Format cell value
# ---------------------------------------------------------------------------

def fmt(value, vtype):
    """Return (display_string, fill_colour) for a cell."""
    if value is None or value == 0:
        return ("", None)

    if vtype in ("pct", "stub_pct"):
        pct = value * 100
        s = f"{pct:+.1f}%"
        colour = C_POS if value > 0 else C_NEG
        if vtype == "stub_pct":
            colour = C_STUB
        return (s, colour)

    if vtype == "flat":
        s = f"{value:+.0f}" if isinstance(value, float) else f"{value:+}"
        colour = C_POS if value > 0 else C_NEG
        return (s, colour)

    if vtype == "cap":
        s = f"{value:+.0f}" if isinstance(value, float) else f"+{value}"
        return (s, C_POS)

    if vtype == "text":
        if isinstance(value, list) and value:
            return (", ".join(value), C_NEG)
        return ("", None)

    return (str(value), None)


# ---------------------------------------------------------------------------
# Write the workbook
# ---------------------------------------------------------------------------

def write_sheet(ws, section_title, rows_iter, wb):
    """Write a single section sheet."""

    # --- Row 1: sheet title ---
    ws.row_dimensions[1].height = 22
    c = ws.cell(1, 1, section_title)
    c.font = bold(C_SHEET_HDR[0:], size=13)  # just use default white text later
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = fill(C_SHEET_HDR)
    c.alignment = left(wrap=False)
    ws.merge_cells(start_row=1, start_column=1, end_row=1,
                   end_column=4 + len(COLUMN_DEFS))

    # --- Row 2: column-group headers (spanning) ---
    ws.row_dimensions[2].height = 18
    # Fixed columns: A=Source Type, B=Name, C=Category/Axis, D=Notes
    for ci, lbl in enumerate(["Source Type", "Name", "Category / Axis", "Notes"], 1):
        c = ws.cell(2, ci, lbl)
        c.font = bold("FFFFFF", size=9)
        c.fill = fill(C_SHEET_HDR)
        c.alignment = center()
        c.border = BORDER

    # Group header spans
    groups_seen = {}
    for i, (grp, key, short, vtype) in enumerate(COLUMN_DEFS):
        col = 5 + i
        groups_seen.setdefault(grp, []).append(col)

    for grp, cols in groups_seen.items():
        start, end = cols[0], cols[-1]
        if start == end:
            c = ws.cell(2, start, grp)
        else:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
            c = ws.cell(2, start, grp)
        c.font = bold("FFFFFF", size=9)
        c.fill = fill(C_GROUP_HDR)
        c.alignment = center()
        c.border = BORDER
        for col in cols[1:]:
            ws.cell(2, col).border = BORDER

    # --- Row 3: individual column headers ---
    ws.row_dimensions[3].height = 44
    for ci, lbl in enumerate(["Source Type", "Name", "Category / Axis", "Notes"], 1):
        c = ws.cell(3, ci, lbl)
        c.font = bold(C_SHEET_HDR, size=9)
        c.fill = fill(C_COL_HDR)
        c.alignment = center()
        c.border = BORDER
    for i, (grp, key, short, vtype) in enumerate(COLUMN_DEFS):
        col = 5 + i
        stub_mark = " *" if "stub" in vtype else ""
        c = ws.cell(3, col, short + stub_mark)
        c.font = bold(C_SHEET_HDR, size=8)
        c.fill = fill(C_COL_HDR)
        c.alignment = center()
        c.border = BORDER

    # --- Data rows ---
    row_num = 4
    for row_data in rows_iter:
        source_type, name, category, note, effects, is_group_start = row_data

        ws.row_dimensions[row_num].height = 15
        row_fill = fill(C_ROW_A) if (row_num % 2 == 0) else fill(C_ROW_B)

        # Category separator styling
        if is_group_start:
            row_fill = fill(C_ROW_CAT)

        for ci, val in enumerate([source_type, name, category, note], 1):
            c = ws.cell(row_num, ci, val)
            c.font = bold(size=9) if is_group_start else std(size=9)
            c.fill = row_fill
            c.alignment = left(wrap=False)
            c.border = BORDER

        for i, (grp, key, short, vtype) in enumerate(COLUMN_DEFS):
            col = 5 + i
            raw_val = effects.get(key)
            display, colour = fmt(raw_val, vtype)
            c = ws.cell(row_num, col, display)
            c.font = std(size=9)
            c.fill = fill(colour) if colour else row_fill
            c.alignment = center(wrap=False)
            c.border = BORDER

        row_num += 1

    # --- Column widths ---
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 36
    for i in range(len(COLUMN_DEFS)):
        col_letter = get_column_letter(5 + i)
        ws.column_dimensions[col_letter].width = 9

    # Freeze panes: lock rows 1-3 and columns A-D
    ws.freeze_panes = "E4"


def write_legend(ws):
    """Write a legend / key sheet."""
    rows = [
        ("Colour Key", ""),
        ("Green cell",  "Positive effect — active in current simulation"),
        ("Red cell",    "Negative effect — active in current simulation"),
        ("Yellow cell", "Stub — declared in constants but not yet wired into simulation"),
        ("Blue row",    "First entry in a new category / axis group"),
        ("", ""),
        ("Column Notes", ""),
        ("stability_recovery_bonus", "Added to the base +0.3/turn province recovery rate. Only active when province is food-sufficient."),
        ("stability (flat)",         "Permanent flat addition to national stability baseline. Sources: government components, traits."),
        ("BE: [category]",           "Building efficiency bonus for that building category. 5 sources stack multiplicatively for government+trait, additively with other sources."),
        ("Columns marked *",         "Stub effect — not yet wired. Declared in constants for future implementation."),
        ("", ""),
        ("Scale Notes", ""),
        ("Buildings",   "Effects shown are base values at Level 1. Scale formula: value × (1 + 0.9 × log₁₀(N)) for level N."),
        ("Government",  "Per-component values. Final national effect = multiplicative product of all 5 chosen components (except stability/growth which are additive)."),
        ("Traits",      "Strong and Weak rows shown separately. A nation picks 1 strong + 2 weak from 3 different pairs."),
    ]
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90

    for i, (key, val) in enumerate(rows, 1):
        ws.row_dimensions[i].height = 16
        ca = ws.cell(i, 1, key)
        cb = ws.cell(i, 2, val)
        if key in ("Colour Key", "Column Notes", "Scale Notes"):
            ca.font = bold(size=10)
            ca.fill = fill(C_SHEET_HDR)
            ca.font = Font(bold=True, size=10, color="FFFFFF")
        elif key.endswith("cell") or key == "Blue row":
            colour_map = {
                "Green cell":  C_POS,
                "Red cell":    C_NEG,
                "Yellow cell": C_STUB,
                "Blue row":    C_ROW_CAT,
            }
            ca.fill = fill(colour_map[key])
            ca.font = std(size=9)
        else:
            ca.font = std(size=9)
        cb.font = std(size=9)
        cb.alignment = left()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Buildings ----
    ws_b = wb.active
    ws_b.title = "Buildings"
    write_sheet(ws_b, "Buildings — base effects at Level 1", building_rows(), wb)

    # ---- Sheet 2: Government ----
    ws_g = wb.create_sheet("Government Options")
    write_sheet(ws_g, "Government — per-component effects (five axes)", government_rows(), wb)

    # ---- Sheet 3: Traits ----
    ws_t = wb.create_sheet("Traits")
    write_sheet(ws_t, "Ideology Traits — strong and weak effects", trait_rows(), wb)

    # ---- Sheet 4: Legend ----
    ws_l = wb.create_sheet("Legend")
    write_legend(ws_l)

    out_path = os.path.join(os.path.dirname(__file__), "effects_matrix.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
