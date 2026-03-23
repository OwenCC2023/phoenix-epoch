"""
Regenerate the "Policy Effects" sheet in effects_matrix.xlsx.

Run from the repo root:
    cd backend
    DJANGO_SETTINGS_MODULE=phoenix_epoch.settings.dev ./venv/Scripts/python.exe ../tools/export_policy_effects.py

Requires openpyxl:
    ./venv/Scripts/pip.exe install openpyxl
"""

import sys, os
sys.stdout.reconfigure(encoding="utf-8")

# Allow running from backend/ with repo root on path
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

import django
django.setup()

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from nations.policy_constants import POLICY_CATEGORIES, POLICY_EFFECTS

# ---------------------------------------------------------------------------
# Colours (match existing sheets)
# ---------------------------------------------------------------------------
FILL_TITLE      = PatternFill("solid", fgColor="1F3864")   # dark blue  row 1
FILL_GROUP_HDR  = PatternFill("solid", fgColor="2E75B6")   # mid blue   row 2
FILL_COL_HDR    = PatternFill("solid", fgColor="BDD7EE")   # light blue row 3
FILL_CAT_FIRST  = PatternFill("solid", fgColor="D6E4F7")   # first row of each category
FILL_LABEL      = PatternFill("solid", fgColor="F9F9F9")   # subsequent label cells
FILL_EMPTY_BLUE = PatternFill("solid", fgColor="D6E4F7")   # empty cells in blue rows
FILL_EMPTY      = PatternFill("solid", fgColor="FFFFFF")   # empty cells elsewhere
FILL_GREEN      = PatternFill("solid", fgColor="C6EFCE")   # positive wired effect
FILL_RED        = PatternFill("solid", fgColor="FFC7CE")   # negative wired effect
FILL_YELLOW     = PatternFill("solid", fgColor="FFF2CC")   # stub effect

FONT_TITLE   = Font(bold=True, color="FFFFFF", size=10)
FONT_GRP_HDR = Font(bold=True, color="FFFFFF", size=8)
FONT_COL_HDR = Font(bold=True, size=8)
FONT_DATA    = Font(size=8)

ALIGN_WRAP   = Alignment(wrap_text=True, vertical="top",    horizontal="left")
ALIGN_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")

MAX_COL = 84

# ---------------------------------------------------------------------------
# Column mapping
# (effect_key) -> (col_number, is_stub, is_percent)
#
# _penalty keys already carry negative values in the constants — no extra
# negation is needed.  The sign of the stored value determines the cell colour.
# ---------------------------------------------------------------------------
EFFECT_COL = {
    "stability_bonus":           (31, False, False),
    "stability_penalty":         (31, False, False),
    "growth_bonus":              (32, False, True),
    "growth_penalty":            (32, False, True),
    "integration_bonus":         (33, False, True),
    "research_bonus":            (35, False, True),
    "research_penalty":          (35, False, True),
    "food_production_bonus":     (38, False, True),
    "wealth_production_bonus":   (40, False, True),
    "manpower_bonus":            (42, False, True),
    "upkeep_reduction":          (20, False, True),
    # Military stubs (yellow)
    "army_training_speed_bonus": (22, True,  True),
    "army_upkeep_reduction":     (24, True,  True),
    "navy_training_speed_bonus": (25, True,  True),
    "navy_upkeep_reduction":     (27, True,  True),
    "air_training_speed_bonus":  (28, True,  True),
    "air_upkeep_reduction":      (30, True,  True),
}

# building_efficiency_bonus sub-keys -> column
BE_CATEGORY_COL = {
    "financial": 43, "light_manufacturing": 44, "heavy_manufacturing": 45,
    "refining": 46, "chemical": 47, "pharmaceutical": 48, "farming": 49,
    "extraction": 50, "construction": 51, "transport": 52, "communications": 53,
    "entertainment": 54, "healthcare": 55, "religious": 56, "green_energy": 57,
    "government_regulatory": 58, "government_oversight": 59,
    "government_management": 60, "government_security": 61,
    "government_education": 62, "government_organization": 63,
    "government_welfare": 64, "military_education": 65,
}


def _fmt(val, is_percent):
    if is_percent:
        return "{:+.1f}%".format(val * 100)
    return "{:+.1f}".format(val)


def _num_fmt(val, is_percent):
    """Return the Excel number_format string for a numeric cell."""
    if is_percent:
        return "+0.0%;-0.0%;0.0%"
    if isinstance(val, int) or (isinstance(val, float) and val == int(val)):
        return "+#,##0;-#,##0;0"
    return "+0.0;-0.0;0.0"


def _write_numeric(cell, val, is_percent, fill):
    cell.value = val
    cell.number_format = _num_fmt(val, is_percent)
    cell.fill = fill
    cell.font = FONT_DATA
    cell.alignment = ALIGN_WRAP


def build_sheet(wb):
    if "Policy Effects" in wb.sheetnames:
        del wb["Policy Effects"]

    ws = wb.create_sheet("Policy Effects")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 24
    for col in range(5, MAX_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 9

    ws.freeze_panes = "E4"

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=MAX_COL)
    c = ws.cell(1, 1, "Policy Effects -- base, government modifiers, and trait modifiers per policy level")
    c.fill = FILL_TITLE; c.font = FONT_TITLE; c.alignment = ALIGN_WRAP

    # Row 2: group headers
    groups = [
        (1,  4,  "Source Type"),
        (5,  15, "Province Effects"),
        (16, 21, "National Capacity"),
        (22, 30, "Military (stub)"),
        (31, 42, "Gov & Trait Modifiers"),
        (43, 65, "Building Efficiency"),
        (66, 76, "Trait Effects"),
        (77, 84, "Stubs"),
    ]
    for start, end, label in groups:
        if start != end:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        c = ws.cell(2, start, label)
        c.fill = FILL_GROUP_HDR; c.font = FONT_GRP_HDR; c.alignment = ALIGN_CENTER

    # Row 3: column headers — copy from Government Options reference sheet
    ws_ref = wb["Government Options"]
    for col in range(1, MAX_COL + 1):
        src = ws_ref.cell(3, col)
        dst = ws.cell(3, col)
        dst.value = src.value
        dst.fill = FILL_COL_HDR; dst.font = FONT_COL_HDR; dst.alignment = ALIGN_WRAP

    # Data rows: one row per (category, level, source_type)
    row = 4
    for cat_key, cat_def in POLICY_CATEGORIES.items():
        cat_name     = cat_def["name"]
        levels_def   = cat_def["levels"]
        cat_effects  = POLICY_EFFECTS.get(cat_key, {})
        first_in_cat = True

        for level_idx, level_def in enumerate(levels_def):
            level_name    = level_def.get("name", "Level " + str(level_idx))
            level_effects = cat_effects.get(level_idx, {})
            if not isinstance(level_effects, dict):
                continue

            base       = level_effects.get("base", {})
            gov_mods   = level_effects.get("government_modifiers", {})
            trait_mods = level_effects.get("trait_modifiers", {})

            sources = []
            if base:
                sources.append(("Policy Base", "base", base))
            for axis_val, mods in gov_mods.items():
                if mods:
                    sources.append(("Policy Gov Mod", "when gov: " + axis_val, mods))
            for trait_key, mods in trait_mods.items():
                if mods:
                    sources.append(("Policy Trait Mod", "when trait: " + trait_key, mods))

            if not sources:
                continue

            for source_label, notes_label, effects in sources:
                is_blue    = first_in_cat
                label_fill = FILL_CAT_FIRST if is_blue else FILL_LABEL

                for col, val in enumerate(
                    [source_label, cat_name + " -- " + level_name, cat_name, notes_label], 1
                ):
                    c = ws.cell(row, col, val)
                    c.fill = label_fill; c.font = FONT_DATA; c.alignment = ALIGN_WRAP

                for col in range(5, MAX_COL + 1):
                    c = ws.cell(row, col)
                    c.fill = FILL_EMPTY_BLUE if is_blue else FILL_EMPTY
                    c.font = FONT_DATA; c.alignment = ALIGN_WRAP

                for key, val in effects.items():
                    if key == "building_efficiency_bonus" and isinstance(val, dict):
                        for be_cat, be_bonus in val.items():
                            col = BE_CATEGORY_COL.get(be_cat)
                            if col:
                                _write_numeric(
                                    ws.cell(row, col), be_bonus, True,
                                    FILL_GREEN if be_bonus >= 0 else FILL_RED,
                                )
                    elif key in EFFECT_COL:
                        col_num, is_stub, is_percent = EFFECT_COL[key]
                        fill = FILL_YELLOW if is_stub else (FILL_GREEN if val >= 0 else FILL_RED)
                        _write_numeric(ws.cell(row, col_num), val, is_percent, fill)

                row += 1
                first_in_cat = False

    print("Policy Effects rows written:", row - 4)

    # Place after Traits
    sheets = wb.sheetnames
    target  = sheets.index("Traits") + 1
    current = sheets.index("Policy Effects")
    wb.move_sheet("Policy Effects", offset=target - current)


if __name__ == "__main__":
    wb_path = os.path.join(ROOT, "effects_matrix.xlsx")
    wb = openpyxl.load_workbook(wb_path)
    build_sheet(wb)
    wb.save(wb_path)
    print("Saved:", wb_path)
