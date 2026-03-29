"""
Import data from policy_effects_complete.xlsx and generate Python source files.

Run from the repo root:
    ./backend/venv/Scripts/python.exe tools/import_from_excel.py

Generates:
    backend/nations/policy_effects_data.py
    backend/nations/gov_policy_multipliers.py
    backend/nations/disabling_rules.py
    backend/nations/policy_building_forbidden.py

Also prints update dicts for government_constants.py and trait_constants.py
that need to be manually integrated (since those files have structural code
beyond just data).
"""

import sys
import os
import json
import textwrap

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with:")
    print("  ./backend/venv/Scripts/pip.exe install openpyxl")
    sys.exit(1)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(ROOT, "policy_effects_complete.xlsx")
BACKEND = os.path.join(ROOT, "backend")
NATIONS = os.path.join(BACKEND, "nations")

# ---------------------------------------------------------------------------
# Column mappings: Excel column index (0-based) -> Python effect key
# ---------------------------------------------------------------------------

# Province effects (cols 2-12)
PROVINCE_EFFECT_COLS = {
    2: "stability_recovery_bonus",
    3: "growth_bonus",         # province-scope growth/turn
    4: "farming_bonus",
    5: "research_bonus",       # province-scope research bonus
    6: "integration_bonus",
    7: "construction_time_reduction",
    8: "literacy_bonus",
    9: "march_speed_bonus",
    10: "sea_transit_speed",
    11: "river_transit_speed",
    12: "air_transit_speed",
}

# National capacity (cols 13-18)
NATIONAL_CAPACITY_COLS = {
    13: "land_trade_capacity",
    14: "naval_trade_capacity",
    15: "air_trade_capacity",
    16: "bureaucratic_capacity",
    17: "upkeep_reduction",
    18: "construction_cost_reduction",
}

# Military stubs (cols 19-27)
MILITARY_COLS = {
    19: "army_training_speed_bonus",
    20: "army_combat_bonus",
    21: "army_upkeep_reduction",
    22: "navy_training_speed_bonus",
    23: "navy_combat_bonus",
    24: "navy_upkeep_reduction",
    25: "air_training_speed_bonus",
    26: "air_combat_bonus",
    27: "air_upkeep_reduction",
}

# Gov & trait modifiers (cols 28-39)
GOV_TRAIT_MODIFIER_COLS = {
    28: "stability_bonus",       # flat
    29: "growth_rate",           # national growth/turn
    30: "integration_pct",
    31: "trade_pct",
    32: "research_pct",
    33: "military_pct",
    34: "consumption_pct",
    35: "production_food_pct",
    36: "production_materials_pct",
    37: "production_wealth_pct",
    38: "production_energy_pct",
    39: "production_manpower_pct",
}

# Building efficiency (cols 40-62) -> category key
BE_COLS = {
    40: "financial",
    41: "light_manufacturing",
    42: "heavy_manufacturing",
    43: "refining",
    44: "chemical",
    45: "pharmaceutical",
    46: "farming",
    47: "extraction",
    48: "construction",
    49: "transport",
    50: "communications",
    51: "entertainment",
    52: "healthcare",
    53: "religious",
    54: "green_energy",
    55: "government_regulatory",
    56: "government_oversight",
    57: "government_management",
    58: "government_security",
    59: "government_education",
    60: "government_organization",
    61: "government_welfare",
    62: "military_education",
}

# Trait-specific effects (cols 63-79)
TRAIT_EFFECT_COLS = {
    63: "manpower_bonus",
    64: "wealth_production_bonus",
    65: "food_production_bonus",
    66: "rural_output_bonus",
    67: "rural_output_penalty",
    68: "urban_output_bonus",
    69: "urban_growth_penalty",
    70: "urban_threshold_reduction",
    71: "training_speed_bonus",
    72: "military_upkeep_reduction",
    73: "building_restrictions",  # special: comma-separated string
    74: "trade_capacity_bonus",
    75: "trade_capacity_penalty",
    76: "espionage_bonus",
    77: "counter_espionage_bonus",
    78: "arms_production_bonus",
    79: "arms_production_penalty",
}

# Building forbidden (cols 80-158) -> building label -> building_type key
# This mapping is built dynamically from the Excel header row and BUILDING_LABEL_TO_KEY

# New effects (cols 159-161)
NEW_EFFECT_COLS = {
    159: "corruption_resistance",
    160: "environmental_health",
    161: "worker_productivity",
}

# ALL numeric effect columns (excluding building_forbidden and building_restrictions)
ALL_EFFECT_COLS = {}
ALL_EFFECT_COLS.update(PROVINCE_EFFECT_COLS)
ALL_EFFECT_COLS.update(NATIONAL_CAPACITY_COLS)
ALL_EFFECT_COLS.update(MILITARY_COLS)
ALL_EFFECT_COLS.update(GOV_TRAIT_MODIFIER_COLS)
ALL_EFFECT_COLS.update(TRAIT_EFFECT_COLS)
ALL_EFFECT_COLS.update(NEW_EFFECT_COLS)
# Remove non-numeric keys
ALL_EFFECT_COLS.pop(73, None)  # building_restrictions is string, handled separately

# Building label -> building_type key mapping
BUILDING_LABEL_TO_KEY = {
    "Bank": "bank",
    "Stock Exchange": "stock_exchange",
    "Trading Post": "trading_post",
    "Electronics Factory": "electronics_factory",
    "Factory": "factory",
    "Precision Workshop": "precision_workshop",
    "Textile Mill": "textile_mill",
    "Arms Factory": "arms_factory",
    "Heavy Forge": "heavy_forge",
    "Industrial Complex": "industrial_complex",
    "Shipyard": "shipyard",
    "Weapons Factory": "weapons_factory",
    "Workshop": "workshop",
    "Advanced Refinery": "advanced_refinery",
    "Biofuel Plant": "biofuel_plant",
    "Fuel Depot": "fuel_depot",
    "Refinery": "refinery",
    "Chemical Plant": "chemical_plant",
    "Fertilizer Plant": "fertilizer_plant",
    "Plastics Factory": "plastics_factory",
    "Medical Supply Depot": "medical_supply_depot",
    "Pharmaceutical Lab": "pharmaceutical_lab",
    "Research Institute": "research_institute",
    "Agricultural Station": "agricultural_station",
    "Grain Silo": "grain_silo",
    "Irrigation Network": "irrigation_network",
    "Logging Camp": "logging_camp",
    "Mine": "mine",
    "Oil Well": "oil_well",
    "Cement Plant": "cement_plant",
    "Construction Yard": "construction_yard",
    "Infrastructure Bureau": "infrastructure_bureau",
    "Air Cargo Terminal": "air_cargo_terminal",
    "Airport": "airport",
    "Bridge": "bridge",
    "Dock": "dock",
    "Logistics Hub": "logistics_hub",
    "Port": "port",
    "Railroad": "railroad",
    "Railway Station": "railway_station",
    "Road Network": "road_network",
    "Train Cargo Terminal": "train_cargo_terminal",
    "Train Depot": "train_depot",
    "Train Station": "train_station",
    "Broadcasting Station": "broadcasting_station",
    "Radio Tower": "radio_tower",
    "Telegraph Network": "telegraph_network",
    "Resort": "resort",
    "Tavern": "tavern",
    "Theatre": "theatre",
    "Clinic": "clinic",
    "Hospital": "hospital",
    "Sanitation Works": "sanitation_works",
    "Church": "church",
    "Holy Site": "holy_site",
    "Madrasa": "madrasa",
    "Hydroelectric Dam": "hydroelectric_dam",
    "Solar Array": "solar_array",
    "Wind Farm": "wind_farm",
    "Regulatory Office": "regulatory_office",
    "Standards Bureau": "standards_bureau",
    "Audit Commission": "audit_commission",
    "Inspector General": "inspector_general",
    "Administrative Center": "administrative_center",
    "Civil Service Academy": "civil_service_academy",
    "Intelligence Agency": "intelligence_agency",
    "Police Headquarters": "police_headquarters",
    "Public School": "public_school",
    "University": "university",
    "Labor Bureau": "labor_bureau",
    "Workers' Council": "workers_council",
    "Public Housing": "public_housing",
    "Social Services Office": "social_services_office",
    "Air Force Academy": "air_force_academy",
    "Military Academy": "military_academy",
    "Naval War College": "naval_war_college",
    "Air Base": "air_base",
    "Army Base": "army_base",
    "Naval Base": "naval_base",
}

# Gov option Excel name -> Python key
GOV_OPTION_NAME_TO_KEY = {
    "Top-Down": "top_down",
    "Bottom-Up": "bottom_up",
    "None": "none",
    "Liberal": "liberal",
    "Collectivist": "collectivist",
    "Protectionist": "protectionist",
    "Resource": "resource",
    "Autarkic": "autarkic",
    "Subsistence": "subsistence",
    "Hereditary": "hereditary",
    "Power-Consensus": "power_consensus",
    "Federal": "federal",
    "Representative": "representative",
    "Direct": "direct",
    "Elections": "elections",
    "Economic Success": "economic_success",
    "Law & Order": "law_and_order",
    "Military Power": "military_power",
    "Religious": "religious",
    "Ideology": "ideology",
    "Singular": "singular",
    "Council": "council",
    "Large Body": "large_body",
    "Multi-body": "multi_body",
    "Staggered Groups": "staggered_groups",
}

GOV_AXIS_NAME_TO_KEY = {
    "Direction": "direction",
    "Economic Cat.": "economic_category",
    "Economic Category": "economic_category",
    "Structure": "structure",
    "Power Origin": "power_origin",
    "Power Type": "power_type",
}


def _safe_float(val):
    """Convert Excel cell value to float, or None if empty."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _slugify_category(name):
    """Convert a policy category display name to a Python key."""
    slug = name.lower().strip()
    slug = slug.replace("&", "and")
    slug = slug.replace("'", "")
    # Replace non-alphanumeric with underscore
    result = []
    for ch in slug:
        if ch.isalnum():
            result.append(ch)
        else:
            result.append("_")
    slug = "".join(result)
    # Collapse consecutive underscores
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_")


def _extract_effects_from_row(vals, col_offset=0):
    """Extract numeric effects from a row using ALL_EFFECT_COLS + BE_COLS.

    col_offset: added to all column indices. The Policy Effects sheet has
    effects starting at col 2 (offset=0), while Government Options and
    Traits Effects sheets have 2 extra header columns so effects start at
    col 4 (offset=2).
    """
    effects = {}

    # Standard numeric effects
    for col_idx, key in ALL_EFFECT_COLS.items():
        actual_col = col_idx + col_offset
        if actual_col < len(vals):
            v = _safe_float(vals[actual_col])
            if v is not None and v != 0:
                effects[key] = v

    # Building efficiency (nested dict)
    be = {}
    for col_idx, cat_key in BE_COLS.items():
        actual_col = col_idx + col_offset
        if actual_col < len(vals):
            v = _safe_float(vals[actual_col])
            if v is not None and v != 0:
                be[cat_key] = v
    if be:
        effects["building_efficiency"] = be

    # Building restrictions (string, for traits)
    restrictions_col = 73 + col_offset
    if restrictions_col < len(vals) and vals[restrictions_col]:
        restrictions = [s.strip() for s in str(vals[restrictions_col]).split(",") if s.strip()]
        if restrictions:
            effects["building_restrictions"] = restrictions

    return effects


def _extract_building_forbidden(vals, building_col_map):
    """Extract building forbidden flags from a policy row."""
    forbidden = set()
    for col_idx, bldg_key in building_col_map.items():
        if col_idx < len(vals):
            v = vals[col_idx]
            if v is not None and v == 1:
                forbidden.add(bldg_key)
    return forbidden


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

def read_government_options(wb):
    """Read Government Options sheet -> dict of gov_key -> effects dict."""
    ws = wb["Government Options"]

    # Build header-based column map from row 3 (Gov Options has headers in row 3)
    header_row = [c.value for c in ws[3]]
    col_map = _build_header_col_map(header_row)

    results = {}

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        vals = list(row)
        source_type = vals[0]
        name = vals[1]
        axis = vals[2]

        if not name or source_type != "Government":
            continue

        key = GOV_OPTION_NAME_TO_KEY.get(str(name).strip())
        if not key:
            print(f"WARNING: Unknown gov option name: {name}")
            continue

        effects = _extract_effects_by_header(vals, col_map)
        results[key] = {
            "name": str(name).strip(),
            "axis": str(axis).strip() if axis else "",
            "effects": effects,
        }

    return results


def _build_header_col_map(header_row):
    """Build a mapping from header display names to effect keys.

    The Traits Effects and Government Options sheets have different column
    layouts (extra columns like Diplo Rep +/-), so we match by header name
    rather than fixed column index.
    """
    HEADER_TO_KEY = {
        "+Stability\nRecovery/turn": "stability_recovery_bonus",
        "+Growth\n/turn": "growth_bonus",
        "+Farming\nBonus": "farming_bonus",
        "+Research\nBonus": "research_bonus",
        "+Integration\nBonus": "integration_bonus",
        "-Construction\nTime": "construction_time_reduction",
        "\u2212Construction\nTime": "construction_time_reduction",  # minus sign variant
        "+Literacy\nBonus": "literacy_bonus",
        "+March\nSpeed": "march_speed_bonus",
        "+Sea\nTransit": "sea_transit_speed",
        "+River\nTransit": "river_transit_speed",
        "+Air\nTransit": "air_transit_speed",
        "Land Trade\nCapacity": "land_trade_capacity",
        "Naval Trade\nCapacity": "naval_trade_capacity",
        "Air Trade\nCapacity": "air_trade_capacity",
        "Bureaucratic\nCapacity": "bureaucratic_capacity",
        "-Upkeep": "upkeep_reduction",
        "\u2212Upkeep": "upkeep_reduction",
        "-Build\nCost": "construction_cost_reduction",
        "\u2212Build\nCost": "construction_cost_reduction",
        "+Army\nTraining *": "army_training_speed_bonus",
        "+Army\nCombat *": "army_combat_bonus",
        "-Army\nUpkeep *": "army_upkeep_reduction",
        "\u2212Army\nUpkeep *": "army_upkeep_reduction",
        "+Navy\nTraining *": "navy_training_speed_bonus",
        "+Navy\nCombat *": "navy_combat_bonus",
        "-Navy\nUpkeep *": "navy_upkeep_reduction",
        "\u2212Navy\nUpkeep *": "navy_upkeep_reduction",
        "+Air\nTraining *": "air_training_speed_bonus",
        "+Air\nCombat *": "air_combat_bonus",
        "-Air\nUpkeep *": "air_upkeep_reduction",
        "\u2212Air\nUpkeep *": "air_upkeep_reduction",
        "Stability\n(flat)": "stability_bonus",
        "Growth\n/turn": "growth_rate",
        "Integration\n(%)": "integration_pct",
        "Trade\n(%)": "trade_pct",
        "Research\n(%)": "research_pct",
        "Military\n(%)": "military_pct",
        "Consumption\n(%)": "consumption_pct",
        "Production\nFood (%)": "production_food_pct",
        "Prod\nFood (%)": "production_food_pct",
        "Production\nMaterials (%)": "production_materials_pct",
        "Prod\nMaterials (%)": "production_materials_pct",
        "Production\nWealth (%)": "production_wealth_pct",
        "Prod\nWealth (%)": "production_wealth_pct",
        "Production\nEnergy (%)": "production_energy_pct",
        "Prod\nEnergy (%)": "production_energy_pct",
        "Production\nManpower (%)": "production_manpower_pct",
        "Prod\nManpower (%)": "production_manpower_pct",
        "+Manpower\n(%)": "manpower_bonus",
        "+Wealth\nProd (%)": "wealth_production_bonus",
        "+Food\nProd (%)": "food_production_bonus",
        "+Rural\nOutput": "rural_output_bonus",
        "-Rural\nOutput": "rural_output_penalty",
        "\u2212Rural\nOutput": "rural_output_penalty",
        "+Urban\nOutput": "urban_output_bonus",
        "-Urban\nGrowth/turn": "urban_growth_penalty",
        "\u2212Urban\nGrowth/turn": "urban_growth_penalty",
        "Urban\nThreshold -": "urban_threshold_reduction",
        "Urban\nThreshold \u2212": "urban_threshold_reduction",
        "+Training\nSpeed *": "training_speed_bonus",
        "-Mil.\nUpkeep *": "military_upkeep_reduction",
        "\u2212Mil.\nUpkeep *": "military_upkeep_reduction",
        "Building\nRestrictions": "building_restrictions",
        "Trade Cap\nBonus *": "trade_capacity_bonus",
        "Trade Cap\nBonus (stub) *": "trade_capacity_bonus",
        "Trade Cap\nPenalty *": "trade_capacity_penalty",
        "Trade Cap\nPenalty (stub) *": "trade_capacity_penalty",
        "Diplo Rep\n+ *": "diplomatic_reputation_bonus",
        "Diplo Rep\n- *": "diplomatic_reputation_penalty",
        "Espionage\n*": "espionage_bonus",
        "Espionage\n(stub) *": "espionage_bonus",
        "Counter\nEspionage *": "counter_espionage_bonus",
        "Counter\nEspionage (stub) *": "counter_espionage_bonus",
        "Arms Prod\nBonus *": "arms_production_bonus",
        "Arms Prod\nBonus (stub) *": "arms_production_bonus",
        "Arms Prod\nPenalty *": "arms_production_penalty",
        "Arms Prod\nPenalty (stub) *": "arms_production_penalty",
        "Corruption\nResistance": "corruption_resistance",
        "Environmental\nHealth": "environmental_health",
        "Worker\nProductivity": "worker_productivity",
    }

    # BE headers use patterns like "BE: Financial", "BE: Light\nManuf."
    BE_HEADER_TO_CAT = {
        "BE:\nFinancial": "financial",
        "BE: Financial": "financial",
        "BE: Light\nManuf.": "light_manufacturing",
        "BE: Heavy\nManuf.": "heavy_manufacturing",
        "BE:\nRefining": "refining",
        "BE:\nChemical": "chemical",
        "BE:\nPharma": "pharmaceutical",
        "BE:\nFarming": "farming",
        "BE:\nExtraction": "extraction",
        "BE:\nConstruction": "construction",
        "BE:\nTransport": "transport",
        "BE:\nComms": "communications",
        "BE:\nEntertainment": "entertainment",
        "BE:\nHealthcare": "healthcare",
        "BE:\nReligious": "religious",
        "BE: Green\nEnergy": "green_energy",
        "BE: Gov\nRegulatory": "government_regulatory",
        "BE: Gov\nOversight": "government_oversight",
        "BE: Gov\nManagement": "government_management",
        "BE: Gov\nSecurity": "government_security",
        "BE: Gov\nEducation": "government_education",
        "BE: Gov\nOrganization": "government_organization",
        "BE: Gov\nWelfare": "government_welfare",
        "BE: Military\nEducation": "military_education",
    }

    col_map = {}  # col_index -> (effect_key, "numeric" | "be" | "restrictions")
    for col_idx, header in enumerate(header_row):
        if not header:
            continue
        h = str(header)
        if h in HEADER_TO_KEY:
            key = HEADER_TO_KEY[h]
            kind = "restrictions" if key == "building_restrictions" else "numeric"
            col_map[col_idx] = (key, kind)
        elif h in BE_HEADER_TO_CAT:
            col_map[col_idx] = (BE_HEADER_TO_CAT[h], "be")

    return col_map


def _extract_effects_by_header(vals, col_map):
    """Extract effects from a row using a header-based column map."""
    effects = {}
    be = {}

    for col_idx, (key, kind) in col_map.items():
        if col_idx >= len(vals) or vals[col_idx] is None:
            continue

        if kind == "numeric":
            v = _safe_float(vals[col_idx])
            if v is not None and v != 0:
                effects[key] = v
        elif kind == "be":
            v = _safe_float(vals[col_idx])
            if v is not None and v != 0:
                be[key] = v
        elif kind == "restrictions":
            restrictions = [s.strip() for s in str(vals[col_idx]).split(",") if s.strip()]
            if restrictions:
                effects["building_restrictions"] = restrictions

    if be:
        effects["building_efficiency"] = be

    return effects


def read_traits(wb):
    """Read Traits Effects sheet -> dict of trait_key -> {strong: {}, weak: {}}."""
    ws = wb["Traits Effects"]

    # Build header-based column map from row 2
    header_row = [c.value for c in ws[2]]
    col_map = _build_header_col_map(header_row)

    results = {}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        vals = list(row)
        source_type = vals[0]
        name = vals[1]
        pair = vals[2]
        strength = vals[3]

        if not name or source_type != "Trait":
            continue

        base_name = str(name).split("(")[0].strip().lower()
        strength_key = str(strength).strip().lower() if strength else "strong"

        if base_name not in results:
            results[base_name] = {"pair": str(pair).strip() if pair else "", "strong_effects": {}, "weak_effects": {}}

        effects = _extract_effects_by_header(vals, col_map)

        if strength_key == "strong":
            results[base_name]["strong_effects"] = effects
        else:
            results[base_name]["weak_effects"] = effects

    return results


def read_policy_effects(wb):
    """Read Corrected Policy Effects sheet.

    Returns:
        categories: OrderedDict of category_key -> {name, levels: [{key, name, effects}]}
        building_forbidden: dict of (category_key, level_idx) -> set of building_type keys
    """
    ws = wb["Corrected Policy Effects"]

    # Build building column map from header row
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    building_col_map = {}
    for col_idx in range(80, min(159, len(header_row))):
        label = header_row[col_idx]
        if label:
            label_str = str(label).strip()
            bldg_key = BUILDING_LABEL_TO_KEY.get(label_str)
            if bldg_key:
                building_col_map[col_idx] = bldg_key
            else:
                print(f"WARNING: Unknown building label in col {col_idx}: {label_str}")

    categories = {}
    building_forbidden = {}
    current_cat_name = None
    current_cat_key = None
    level_idx = 0

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        vals = list(row)
        cat_name = vals[0]
        level_name = vals[1]

        if not level_name:
            continue

        level_name = str(level_name).strip()

        # New category
        if cat_name and str(cat_name).strip():
            current_cat_name = str(cat_name).strip()
            current_cat_key = _slugify_category(current_cat_name)
            level_idx = 0
            categories[current_cat_key] = {
                "name": current_cat_name,
                "levels": [],
            }
        else:
            level_idx += 1

        if not current_cat_key:
            continue

        # Extract effects
        effects = _extract_effects_from_row(vals)

        # Extract building forbidden
        forbidden = _extract_building_forbidden(vals, building_col_map)

        level_key = _slugify_category(level_name)
        categories[current_cat_key]["levels"].append({
            "key": level_key,
            "name": level_name,
            "effects": effects,
        })

        if forbidden:
            building_forbidden[(current_cat_key, level_idx)] = forbidden

    return categories, building_forbidden


def read_gov_policy_multipliers(wb):
    """Read Gov-Policy Multipliers sheet -> dict of gov_key -> {(cat_key, level_idx): multiplier}."""
    ws = wb["Gov-Policy Multipliers"]

    # Row 1 has category names, row 2 has level names
    row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    # Build column -> (category_key, level_idx) mapping from headers
    col_to_policy = {}
    current_cat = None
    current_cat_key = None
    level_idx = 0

    for col_idx in range(2, len(row2)):
        cat_name = row1[col_idx] if col_idx < len(row1) else None
        level_name = row2[col_idx] if col_idx < len(row2) else None

        if not level_name:
            continue

        if cat_name and str(cat_name).strip():
            current_cat = str(cat_name).strip()
            current_cat_key = _slugify_category(current_cat)
            level_idx = 0
        else:
            level_idx += 1

        if current_cat_key:
            col_to_policy[col_idx] = (current_cat_key, level_idx)

    # Read gov option rows (starting row 3)
    results = {}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        vals = list(row)
        axis = vals[0]
        name = vals[1]

        if not name:
            continue

        gov_key = GOV_OPTION_NAME_TO_KEY.get(str(name).strip())
        if not gov_key:
            print(f"WARNING: Unknown gov option in multipliers: {name}")
            continue

        multipliers = {}
        for col_idx, policy_id in col_to_policy.items():
            if col_idx < len(vals):
                v = _safe_float(vals[col_idx])
                if v is not None and v != 1.0:
                    multipliers[policy_id] = v

        if multipliers:
            results[gov_key] = multipliers

    return results


def read_gov_policy_disabling(wb):
    """Read Gov-Policy Disabling sheet -> dict of gov_key -> [(cat_key, level_name)]."""
    ws = wb["Gov-Policy Disabling"]
    results = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals or not vals[0]:
            continue

        gov_axis = str(vals[0]).strip() if vals[0] else ""
        gov_name = str(vals[1]).strip() if vals[1] else ""
        disabled_policy = str(vals[2]).strip() if vals[2] else ""
        disabled_level = str(vals[3]).strip() if vals[3] else ""
        reason = str(vals[4]).strip() if len(vals) > 4 and vals[4] else ""

        gov_key = GOV_OPTION_NAME_TO_KEY.get(gov_name)
        if not gov_key:
            print(f"WARNING: Unknown gov option in disabling: {gov_name}")
            continue

        if gov_key not in results:
            results[gov_key] = []

        results[gov_key].append({
            "policy": disabled_policy,
            "level": disabled_level,
            "reason": reason,
        })

    return results


def read_trait_disabling_rules(wb):
    """Read Trait Disabling Rules sheet -> separate dicts by target type."""
    ws = wb["Trait Disabling Rules"]

    gov_disables = {}    # (trait, strength) -> [gov_option_key]
    policy_disables = {} # (trait, strength) -> [(policy_name, level_name)]
    building_disables = {} # (trait, strength) -> [building_key/name]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals or not vals[0]:
            continue

        trait_name = str(vals[0]).strip().lower() if vals[0] else ""
        strength = str(vals[1]).strip().lower() if vals[1] else "strong"
        target_type = str(vals[2]).strip() if vals[2] else ""
        target_item = str(vals[3]).strip() if vals[3] else ""
        reason = str(vals[4]).strip() if len(vals) > 4 and vals[4] else ""

        key = (trait_name, strength)

        if target_type == "Gov Option":
            if key not in gov_disables:
                gov_disables[key] = []
            gov_disables[key].append({
                "target": target_item,
                "reason": reason,
            })
        elif target_type == "Policy":
            if key not in policy_disables:
                policy_disables[key] = []
            # Parse "Policy: Level" format - target_item might be "Military Service: All Adults Serve"
            policy_disables[key].append({
                "target": target_item,
                "reason": reason,
            })
        elif target_type == "Building":
            if key not in building_disables:
                building_disables[key] = []
            building_disables[key].append({
                "target": target_item,
                "reason": reason,
            })

    return gov_disables, policy_disables, building_disables


def read_policy_disabling_rules(wb):
    """Read Policy Disabling Rules sheet -> list of rules."""
    ws = wb["Policy Disabling Rules"]
    rules = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        vals = list(row)
        if not vals or not vals[0]:
            continue

        policy = str(vals[0]).strip() if vals[0] else ""
        level = str(vals[1]).strip() if vals[1] else ""
        disabled_when = str(vals[2]).strip() if vals[2] else ""
        reason = str(vals[3]).strip() if len(vals) > 3 and vals[3] else ""

        rules.append({
            "policy": policy,
            "level": level,
            "disabled_when": disabled_when,
            "reason": reason,
        })

    return rules


# ---------------------------------------------------------------------------
# Code generators
# ---------------------------------------------------------------------------

def _repr_dict(d, indent=0):
    """Pretty-print a dict as valid Python source."""
    if not d:
        return "{}"
    pad = "    " * indent
    inner_pad = "    " * (indent + 1)
    lines = ["{"]
    for k, v in d.items():
        if isinstance(v, dict):
            lines.append(f"{inner_pad}{k!r}: {_repr_dict(v, indent + 1)},")
        elif isinstance(v, set):
            items = sorted(v)
            lines.append(f"{inner_pad}{k!r}: {{{', '.join(repr(x) for x in items)}}},")
        elif isinstance(v, list):
            if all(isinstance(x, str) for x in v):
                lines.append(f"{inner_pad}{k!r}: {v!r},")
            else:
                lines.append(f"{inner_pad}{k!r}: [")
                for item in v:
                    if isinstance(item, dict):
                        lines.append(f"{inner_pad}    {_repr_dict(item, indent + 2)},")
                    elif isinstance(item, tuple):
                        lines.append(f"{inner_pad}    {item!r},")
                    else:
                        lines.append(f"{inner_pad}    {item!r},")
                lines.append(f"{inner_pad}],")
        elif isinstance(v, float):
            # Clean float representation
            if v == int(v) and abs(v) < 1e10:
                lines.append(f"{inner_pad}{k!r}: {int(v)},")
            else:
                lines.append(f"{inner_pad}{k!r}: {v},")
        else:
            lines.append(f"{inner_pad}{k!r}: {v!r},")
    lines.append(f"{pad}}}")
    return "\n".join(lines)


def generate_policy_effects_data(categories):
    """Generate backend/nations/policy_effects_data.py."""
    lines = [
        '"""',
        "Auto-generated policy effects data from policy_effects_complete.xlsx.",
        "",
        "DO NOT EDIT MANUALLY. Re-generate with:",
        "    ./backend/venv/Scripts/python.exe tools/import_from_excel.py",
        '"""',
        "",
        "",
        "# 63 policy categories with their levels and display names.",
        "# Each category has a unique key, a display name, a default_level (0),",
        "# and a list of levels with key, name, and description.",
        "POLICY_CATEGORIES = {",
    ]

    for cat_key, cat_data in categories.items():
        lines.append(f'    {cat_key!r}: {{')
        lines.append(f'        "name": {cat_data["name"]!r},')
        lines.append(f'        "default_level": 0,')
        lines.append(f'        "levels": [')
        for level in cat_data["levels"]:
            lines.append(f'            {{"key": {level["key"]!r}, "name": {level["name"]!r}, "description": "", "tags": [], "effects": {{}}}},')
        lines.append(f'        ],')
        lines.append(f'    }},')

    lines.append("}")
    lines.append("")
    lines.append("")

    # POLICY_EFFECTS dict
    lines.append("# Effects per policy level.")
    lines.append("# Structure: category_key -> level_index -> {\"base\": {effect_key: value}}")
    lines.append("POLICY_EFFECTS = {")

    for cat_key, cat_data in categories.items():
        lines.append(f'    {cat_key!r}: {{')
        for level_idx, level in enumerate(cat_data["levels"]):
            effects = level.get("effects", {})
            if effects:
                lines.append(f'        {level_idx}: {{')
                lines.append(f'            "base": {{')
                for ek, ev in sorted(effects.items()):
                    if isinstance(ev, dict):
                        lines.append(f'                {ek!r}: {{')
                        for bk, bv in sorted(ev.items()):
                            lines.append(f'                    {bk!r}: {bv},')
                        lines.append(f'                }},')
                    elif isinstance(ev, list):
                        lines.append(f'                {ek!r}: {ev!r},')
                    else:
                        lines.append(f'                {ek!r}: {ev},')
                lines.append(f'            }},')
                lines.append(f'        }},')
            else:
                lines.append(f'        {level_idx}: {{"base": {{}}}},')
        lines.append(f'    }},')

    lines.append("}")
    lines.append("")

    return "\n".join(lines)


def generate_gov_policy_multipliers(multipliers):
    """Generate backend/nations/gov_policy_multipliers.py."""
    lines = [
        '"""',
        "Auto-generated Gov-Policy Multiplier matrix from policy_effects_complete.xlsx.",
        "",
        "DO NOT EDIT MANUALLY. Re-generate with:",
        "    ./backend/venv/Scripts/python.exe tools/import_from_excel.py",
        "",
        "Structure: GOV_POLICY_MULTIPLIERS[gov_option_key][(category_key, level_index)] = float",
        "Only non-1.0 entries are stored. Default is 1.0.",
        '"""',
        "",
        "",
        "GOV_POLICY_MULTIPLIERS = {",
    ]

    for gov_key in sorted(multipliers.keys()):
        mults = multipliers[gov_key]
        lines.append(f'    {gov_key!r}: {{')
        for (cat_key, level_idx), value in sorted(mults.items()):
            lines.append(f'        ({cat_key!r}, {level_idx}): {value},')
        lines.append(f'    }},')

    lines.append("}")
    lines.append("")

    return "\n".join(lines)


def generate_disabling_rules(gov_policy_disabling, trait_gov, trait_policy,
                              trait_building, policy_policy):
    """Generate backend/nations/disabling_rules.py."""
    lines = [
        '"""',
        "Auto-generated disabling rules from policy_effects_complete.xlsx.",
        "",
        "DO NOT EDIT MANUALLY. Re-generate with:",
        "    ./backend/venv/Scripts/python.exe tools/import_from_excel.py",
        "",
        "Disabling cascade:",
        "    Traits -> Gov Options, Policies, Buildings",
        "    Gov Options -> Policies",
        "    Policies -> other Policies, Buildings (via POLICY_BUILDING_FORBIDDEN)",
        '"""',
        "",
        "",
    ]

    # Trait -> Gov disabling
    lines.append("# Trait -> Gov Option disabling (17 rules)")
    lines.append("# Key: (trait_name, strength) -> list of dicts with target gov option and reason")
    lines.append("TRAIT_GOV_DISABLES = {")
    for (trait, strength), items in sorted(trait_gov.items()):
        lines.append(f'    ({trait!r}, {strength!r}): [')
        for item in items:
            lines.append(f'        {{"target": {item["target"]!r}, "reason": {item["reason"]!r}}},')
        lines.append(f'    ],')
    lines.append("}")
    lines.append("")

    # Trait -> Policy disabling
    lines.append("# Trait -> Policy disabling (85 rules)")
    lines.append("# Key: (trait_name, strength) -> list of dicts with target policy:level and reason")
    lines.append("TRAIT_POLICY_DISABLES = {")
    for (trait, strength), items in sorted(trait_policy.items()):
        lines.append(f'    ({trait!r}, {strength!r}): [')
        for item in items:
            lines.append(f'        {{"target": {item["target"]!r}, "reason": {item["reason"]!r}}},')
        lines.append(f'    ],')
    lines.append("}")
    lines.append("")

    # Trait -> Building disabling
    lines.append("# Trait -> Building disabling (23 rules)")
    lines.append("# Key: (trait_name, strength) -> list of dicts with target building and reason")
    lines.append("TRAIT_BUILDING_DISABLES = {")
    for (trait, strength), items in sorted(trait_building.items()):
        lines.append(f'    ({trait!r}, {strength!r}): [')
        for item in items:
            lines.append(f'        {{"target": {item["target"]!r}, "reason": {item["reason"]!r}}},')
        lines.append(f'    ],')
    lines.append("}")
    lines.append("")

    # Gov -> Policy disabling
    lines.append("# Gov Option -> Policy disabling (59 rules)")
    lines.append("# Key: gov_option_key -> list of dicts with disabled policy, level, and reason")
    lines.append("GOV_POLICY_DISABLES = {")
    for gov_key, items in sorted(gov_policy_disabling.items()):
        lines.append(f'    {gov_key!r}: [')
        for item in items:
            lines.append(f'        {{"policy": {item["policy"]!r}, "level": {item["level"]!r}, "reason": {item["reason"]!r}}},')
        lines.append(f'    ],')
    lines.append("}")
    lines.append("")

    # Policy -> Policy disabling
    lines.append("# Policy -> Policy disabling (48 rules)")
    lines.append("# Each rule: the policy at the given level is disabled when the condition is met")
    lines.append("POLICY_POLICY_DISABLES = [")
    for rule in policy_policy:
        lines.append(f'    {{"policy": {rule["policy"]!r}, "level": {rule["level"]!r}, "disabled_when": {rule["disabled_when"]!r}, "reason": {rule["reason"]!r}}},')
    lines.append("]")
    lines.append("")

    return "\n".join(lines)


def generate_policy_building_forbidden(forbidden):
    """Generate backend/nations/policy_building_forbidden.py."""
    lines = [
        '"""',
        "Auto-generated policy -> building forbidden mappings from policy_effects_complete.xlsx.",
        "",
        "DO NOT EDIT MANUALLY. Re-generate with:",
        "    ./backend/venv/Scripts/python.exe tools/import_from_excel.py",
        "",
        "Structure: POLICY_BUILDING_FORBIDDEN[(category_key, level_index)] = set of building_type keys",
        '"""',
        "",
        "",
        "POLICY_BUILDING_FORBIDDEN = {",
    ]

    for (cat_key, level_idx), bldg_set in sorted(forbidden.items()):
        items = sorted(bldg_set)
        lines.append(f'    ({cat_key!r}, {level_idx}): {{')
        for item in items:
            lines.append(f'        {item!r},')
        lines.append(f'    }},')

    lines.append("}")
    lines.append("")

    return "\n".join(lines)


def generate_gov_options_update(gov_data):
    """Print updated gov option dicts for manual integration."""
    lines = [
        "",
        "=" * 70,
        "GOVERNMENT OPTIONS UPDATE",
        "Copy these into backend/nations/government_constants.py",
        "=" * 70,
        "",
    ]

    for key, data in sorted(gov_data.items()):
        effects = data["effects"]
        lines.append(f'    "{key}": {{')

        # Separate flat effects from building_efficiency
        be = effects.pop("building_efficiency", {})

        # Standard effects
        for ek in ["stability_bonus", "growth_rate", "integration_pct", "trade_pct",
                    "research_pct", "military_pct", "consumption_pct"]:
            if ek in effects:
                lines.append(f'        "{ek}": {effects[ek]},')

        # Production dict
        prod_keys = {k: v for k, v in effects.items() if k.startswith("production_")}
        if prod_keys:
            lines.append(f'        "production": {{')
            for pk, pv in sorted(prod_keys.items()):
                short = pk.replace("production_", "").replace("_pct", "")
                lines.append(f'            "{short}": {pv},')
            lines.append(f'        }},')

        # Building efficiency
        if be:
            lines.append(f'        "building_efficiency": {{')
            for bk, bv in sorted(be.items()):
                lines.append(f'            "{bk}": {bv},')
            lines.append(f'        }},')

        lines.append(f'    }},')
        lines.append("")

    return "\n".join(lines)


def generate_traits_update(trait_data):
    """Print updated trait dicts for manual integration."""
    lines = [
        "",
        "=" * 70,
        "TRAITS UPDATE",
        "Copy these into backend/nations/trait_constants.py",
        "=" * 70,
        "",
    ]

    for key, data in sorted(trait_data.items()):
        lines.append(f'    "{key}": {{')
        lines.append(f'        "name": "{key.title()}",')
        lines.append(f'        "description": "",')

        # Parse pair index from pair string
        pair_str = data.get("pair", "")
        pair_idx = 0
        if "Pair" in pair_str:
            try:
                pair_idx = int(pair_str.split("Pair")[1].split(":")[0].strip())
            except (ValueError, IndexError):
                pass
        lines.append(f'        "pair_index": {pair_idx},')

        for strength in ["strong", "weak"]:
            effects = data.get(f"{strength}_effects", {})
            lines.append(f'        "{strength}_effects": {{')
            for ek, ev in sorted(effects.items()):
                if isinstance(ev, dict):
                    lines.append(f'            "{ek}": {{')
                    for bk, bv in sorted(ev.items()):
                        lines.append(f'                "{bk}": {bv},')
                    lines.append(f'            }},')
                elif isinstance(ev, list):
                    lines.append(f'            "{ek}": {ev!r},')
                else:
                    lines.append(f'            "{ek}": {ev},')
            lines.append(f'        }},')

        lines.append(f'    }},')
        lines.append("")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # --- Read all sheets ---
    print("Reading Government Options...")
    gov_data = read_government_options(wb)
    print(f"  Found {len(gov_data)} government options")

    print("Reading Traits Effects...")
    trait_data = read_traits(wb)
    print(f"  Found {len(trait_data)} traits")

    print("Reading Corrected Policy Effects...")
    categories, building_forbidden = read_policy_effects(wb)
    total_levels = sum(len(c["levels"]) for c in categories.values())
    print(f"  Found {len(categories)} categories, {total_levels} total levels")
    print(f"  Found {len(building_forbidden)} policy levels with building forbidden rules")

    print("Reading Gov-Policy Multipliers...")
    multipliers = read_gov_policy_multipliers(wb)
    total_mults = sum(len(m) for m in multipliers.values())
    print(f"  Found {len(multipliers)} gov options with non-default multipliers ({total_mults} entries)")

    print("Reading Gov-Policy Disabling...")
    gov_policy_disabling = read_gov_policy_disabling(wb)
    total_gpd = sum(len(v) for v in gov_policy_disabling.values())
    print(f"  Found {total_gpd} gov-policy disabling rules")

    print("Reading Trait Disabling Rules...")
    trait_gov, trait_policy, trait_building = read_trait_disabling_rules(wb)
    print(f"  Trait->Gov: {sum(len(v) for v in trait_gov.values())} rules")
    print(f"  Trait->Policy: {sum(len(v) for v in trait_policy.values())} rules")
    print(f"  Trait->Building: {sum(len(v) for v in trait_building.values())} rules")

    print("Reading Policy Disabling Rules...")
    policy_policy = read_policy_disabling_rules(wb)
    print(f"  Found {len(policy_policy)} policy-policy disabling rules")

    # --- Generate files ---
    print("\nGenerating files...")

    # 1. Policy effects data
    path = os.path.join(NATIONS, "policy_effects_data.py")
    content = generate_policy_effects_data(categories)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Written: {path}")

    # 2. Gov-Policy Multipliers
    path = os.path.join(NATIONS, "gov_policy_multipliers.py")
    content = generate_gov_policy_multipliers(multipliers)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Written: {path}")

    # 3. Disabling rules
    path = os.path.join(NATIONS, "disabling_rules.py")
    content = generate_disabling_rules(
        gov_policy_disabling, trait_gov, trait_policy, trait_building, policy_policy
    )
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Written: {path}")

    # 4. Policy building forbidden
    path = os.path.join(NATIONS, "policy_building_forbidden.py")
    content = generate_policy_building_forbidden(building_forbidden)
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  Written: {path}")

    # --- Print update instructions for manual files ---
    gov_update = generate_gov_options_update(gov_data)
    print(gov_update)

    traits_update = generate_traits_update(trait_data)
    print(traits_update)

    # --- Validation summary ---
    print("\n" + "=" * 70)
    print("VALIDATION SUMMARY")
    print("=" * 70)

    # Check a few known values
    print("\nSpot checks:")

    # Military Service: Disarmed Nation should have stability_bonus ~0.3
    ms = categories.get("military_service", {}).get("levels", [{}])
    if ms:
        ms0 = ms[0].get("effects", {})
        sb = ms0.get("stability_bonus")
        print(f"  Military Service / Disarmed Nation stability_bonus: {sb} (expected ~0.3)")

    # Policing: Protect and Serve should have positive stability
    pol = categories.get("policing", {}).get("levels", [])
    if len(pol) > 3:
        ps = pol[3].get("effects", {})
        sb = ps.get("stability_bonus")
        print(f"  Policing / Protect and Serve stability_bonus: {sb} (expected ~0.8)")

    # Gov: Top-Down should have integration_pct ~0.08
    td = gov_data.get("top_down", {}).get("effects", {})
    print(f"  Gov Top-Down integration_pct: {td.get('integration_pct')} (expected 0.08)")
    print(f"  Gov Top-Down stability_bonus: {td.get('stability_bonus')} (expected -1)")

    # Trait: Militarist strong should have heavy_manufacturing BE ~0.10
    mil = trait_data.get("militarist", {}).get("strong_effects", {})
    be = mil.get("building_efficiency", {})
    print(f"  Trait Militarist (strong) BE heavy_manufacturing: {be.get('heavy_manufacturing')} (expected 0.10)")

    # Multipliers: Top-Down / Policing levels
    td_mults = multipliers.get("top_down", {})
    for key, val in sorted(td_mults.items()):
        if key[0] == "policing":
            print(f"  Multiplier Top-Down / policing level {key[1]}: {val}")
            break

    print("\nDone! Review the generated files and the update instructions above.")
    print("Next steps:")
    print("  1. Manually update government_constants.py with the gov options above")
    print("  2. Manually update trait_constants.py with the traits above")
    print("  3. Update policy_constants.py to import from policy_effects_data.py")
    print("  4. Wire gov_policy_multipliers.py into policy_effects.py")
    print("  5. Wire disabling_rules.py into policy_effects.py validation")
    print("  6. Wire policy_building_forbidden.py into building block checks")


if __name__ == "__main__":
    main()
